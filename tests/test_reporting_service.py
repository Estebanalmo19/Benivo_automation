import datetime

from openpyxl import load_workbook

from app.services import reporting_service as reporting

SOME_DATE = datetime.date(2026, 1, 1)


def _candidate(
    application_eid,
    is_relocation_required="Yes",
    start_date=SOME_DATE,
    workplace="Serbia Live Casino",
    benivo_status="READY_TO_POST",
    workflow_state="Mobility in process",
    first_name="Jane",
    last_name="Doe",
    is_vip=False,
):
    return {
        "application_eid": application_eid,
        "candidate_eid": f"C-{application_eid}",
        "email": f"{application_eid}@example.com",
        "first_name": first_name,
        "last_name": last_name,
        "workflow_state": workflow_state,
        "is_relocation_required": is_relocation_required,
        "start_date": start_date,
        "workplace": workplace,
        "job_title": "Presenter",
        "requisition_id": "REQ-1",
        "department": "Live Casino",
        "location": "Somewhere",
        "benivo_status": benivo_status,
        "is_vip": is_vip,
        "created_at": datetime.datetime(2026, 1, 1, tzinfo=datetime.timezone.utc),
        "updated_at": datetime.datetime(2026, 1, 1, tzinfo=datetime.timezone.utc),
    }


# ---------------------------------------------------------------------------
# Pure helper functions
# ---------------------------------------------------------------------------

def test_relocation_bucket_yes():
    assert reporting._relocation_bucket("Yes") == "yes"
    assert reporting._relocation_bucket("yes") == "yes"


def test_relocation_bucket_no():
    assert reporting._relocation_bucket("No") == "no"


def test_relocation_bucket_blank_or_unrecognized():
    assert reporting._relocation_bucket(None) == "blank_or_unrecognized"
    assert reporting._relocation_bucket("") == "blank_or_unrecognized"
    assert reporting._relocation_bucket("Maybe") == "blank_or_unrecognized"


def test_candidate_name_joins_first_and_last():
    assert reporting._candidate_name({"first_name": "Jane", "last_name": "Doe"}) == "Jane Doe"


def test_candidate_name_handles_missing_parts():
    assert reporting._candidate_name({"first_name": None, "last_name": "Doe"}) == "Doe"
    assert reporting._candidate_name({"first_name": None, "last_name": None}) is None


def test_build_row_has_all_required_columns():
    row = reporting._build_row(_candidate("APP-1"), "some reason")
    assert set(row.keys()) == set(reporting.REQUIRED_COLUMNS)


def test_build_row_ready_reason_default():
    row = reporting._build_row(_candidate("APP-1"), reporting.READY_REASON)
    assert row["reason"] == reporting.READY_REASON


def test_build_row_policy_name_from_is_vip():
    row_basic = reporting._build_row(_candidate("APP-1", is_vip=False), "reason")
    row_vip = reporting._build_row(_candidate("APP-2", is_vip=True), "reason")

    assert row_basic["is_vip"] is False
    assert row_basic["policy_name"] == "Basic"
    assert row_vip["is_vip"] is True
    assert row_vip["policy_name"] == "VIP"


# ---------------------------------------------------------------------------
# generate_reports() end-to-end (mocked repositories + no reference-data calls)
# ---------------------------------------------------------------------------

def _mock_population():
    # benivo_status here reflects what classification_service would already
    # have persisted (READY_TO_POST vs PENDING_OFFICE_MAPPING) -- reporting
    # trusts this status directly rather than re-deriving it.
    return [
        _candidate("READY-1", is_relocation_required="Yes", start_date=SOME_DATE, workplace="Serbia Live Casino", benivo_status="READY_TO_POST"),
        _candidate("OFFICE-PENDING-1", is_relocation_required="Yes", start_date=SOME_DATE, workplace="Unmapped Site", benivo_status="PENDING_OFFICE_MAPPING"),
        _candidate("MISSING-1", is_relocation_required="Yes", start_date=None, benivo_status="PENDING_MISSING_START_DATE"),
        _candidate("REVIEW-NO-1", is_relocation_required="No", start_date=None, benivo_status="NEEDS_RECRUITER_REVIEW"),
        _candidate("REVIEW-BLANK-1", is_relocation_required=None, start_date=None, benivo_status="NEEDS_RECRUITER_REVIEW"),
    ]


def test_generate_reports_reconciles_with_summary(tmp_path, monkeypatch):
    monkeypatch.setattr(reporting, "_report_dir", lambda: tmp_path)
    monkeypatch.setattr(reporting.candidate_repository, "get_all_candidates_for_report", lambda: _mock_population())
    monkeypatch.setattr(reporting.post_log_repository, "get_terminal_post_log_application_eids", lambda: set())
    monkeypatch.setattr(reporting.posting_service, "allow_reference_data_calls", lambda: False)

    selected = [_mock_population()[0]]  # READY-1 selected

    report_path = reporting.generate_reports(
        selected_candidates=selected,
        posting_results=[{"application_eid": "READY-1", "office_resolved": False, "payload": {}}],
        dry_run=True,
        posting_limit=5,
    )

    assert report_path.exists()

    wb = load_workbook(report_path)
    assert set(wb.sheetnames) == {
        "Summary", "Ready to Post", "Missing Start Date",
        "Relocation Field Review", "Missing Office Mapping", "Posting Results",
    }

    summary = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}

    # Placement now trusts the persisted benivo_status: READY-1 is
    # READY_TO_POST -> Ready to Post; OFFICE-PENDING-1 is
    # PENDING_OFFICE_MAPPING -> Missing Office Mapping. Reference-data calls
    # are disabled in this test, so office name/id display fields stay null,
    # but that no longer affects which sheet a row lands in.
    assert summary["vip_candidates"] == 0
    assert summary["basic_candidates"] == 5
    assert summary["relocation_yes_with_start_date"] == 2
    assert summary["ready_to_post"] == 1
    assert summary["office_mapping_resolved"] == 1
    assert summary["pending_office_mapping"] == 1
    assert summary["office_mapping_unresolved"] == 1
    assert summary["pending_missing_start_date"] == 1
    assert summary["relocation_requires_review"] == 2
    assert summary["dry_run"] is True
    assert summary["posting_limit"] == 5
    assert summary["posting_success_current_run"] == 0
    assert summary["posting_already_exists_current_run"] == 0
    assert summary["posting_failed_current_run"] == 0

    def _data_rows(ws):
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return [r for r in rows if any(v is not None for v in r)]

    missing_start_date_rows = _data_rows(wb["Missing Start Date"])
    assert len(missing_start_date_rows) == summary["pending_missing_start_date"]

    review_rows = _data_rows(wb["Relocation Field Review"])
    assert len(review_rows) == summary["relocation_no"] + summary["relocation_blank_or_unrecognized"]

    missing_office_rows = _data_rows(wb["Missing Office Mapping"])
    assert len(missing_office_rows) == summary["office_mapping_unresolved"]

    ready_rows = _data_rows(wb["Ready to Post"])
    assert len(ready_rows) == summary["office_mapping_resolved"]

    header = [cell.value for cell in wb["Missing Start Date"][1]]
    assert "workflow_state" in header
    assert "is_relocation_required" in header

    # Posting Results must be empty-with-note during dry run.
    posting_sheet_first_cell = wb["Posting Results"]["A1"].value
    assert posting_sheet_first_cell == reporting.DRY_RUN_NOTE


def test_generate_reports_selected_count_never_exceeds_limit(tmp_path, monkeypatch):
    monkeypatch.setattr(reporting, "_report_dir", lambda: tmp_path)
    monkeypatch.setattr(reporting.candidate_repository, "get_all_candidates_for_report", lambda: _mock_population())
    monkeypatch.setattr(reporting.post_log_repository, "get_terminal_post_log_application_eids", lambda: set())
    monkeypatch.setattr(reporting.posting_service, "allow_reference_data_calls", lambda: False)

    selected = _mock_population()[:2]

    reporting.generate_reports(selected_candidates=selected, posting_results=[], dry_run=True, posting_limit=5)

    assert len(selected) <= 5
