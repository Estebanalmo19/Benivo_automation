"""Operational Excel report: Summary + candidate-level evidence for every classification."""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app import config
from app.clients import benivo_client
from app.repositories import candidate_repository, post_log_repository
from app.services import posting_service
from app.services.office_resolution_service import resolve_office
from app.services.policy_service import resolve_policy

logger = logging.getLogger(__name__)

REPORT_FILENAME_PREFIX = "benivo_operational_report"

MOBILITY_WORKFLOW_STATE = "Mobility in process"

REQUIRED_COLUMNS = [
    "application_eid",
    "candidate_eid",
    "candidate_name",
    "email",
    "workflow_state",
    "is_relocation_required",
    "start_date",
    "workplace",
    "resolved_benivo_office_name",
    "resolved_benivo_office_id",
    "job_title",
    "department",
    "location",
    "benivo_status",
    "is_vip",
    "policy_name",
    "reason",
    "selected_for_current_run",
]

READY_REASON = "Candidate meets all current posting requirements"
MISSING_START_DATE_REASON = "Start date is missing while relocation is required"
RELOCATION_NO_REASON = "Relocation is marked as No and requires recruiter confirmation"
RELOCATION_UNRECOGNIZED_REASON = "Relocation value is blank or unrecognized and requires recruiter confirmation"
MISSING_OFFICE_REASON = "No confirmed Benivo office mapping exists for the candidate workplace"
DRY_RUN_NOTE = "No real posting was performed (DRY RUN)."


def _report_dir() -> Path:
    configured = config.report_folder()
    report_dir = Path(configured) if configured else Path(__file__).resolve().parent.parent.parent
    report_dir.mkdir(parents=True, exist_ok=True)
    return report_dir


def _excel_safe(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _candidate_name(candidate: Dict[str, Any]) -> Optional[str]:
    parts = [candidate.get("first_name"), candidate.get("last_name")]
    joined = " ".join(part for part in parts if part)
    return joined or None


def _relocation_bucket(is_relocation_required: Optional[str]) -> str:
    value = (is_relocation_required or "").strip().lower()

    if value == "yes":
        return "yes"

    if value == "no":
        return "no"

    return "blank_or_unrecognized"


def _build_row(
    candidate: Dict[str, Any],
    reason: str,
    office: Optional[Dict[str, str]] = None,
    selected: bool = False,
) -> Dict[str, Any]:
    return {
        "application_eid": candidate.get("application_eid"),
        "candidate_eid": candidate.get("candidate_eid"),
        "candidate_name": _candidate_name(candidate),
        "email": candidate.get("email"),
        "workflow_state": candidate.get("workflow_state"),
        "is_relocation_required": candidate.get("is_relocation_required"),
        "start_date": _excel_safe(candidate.get("start_date")),
        "workplace": candidate.get("workplace"),
        "resolved_benivo_office_name": office.get("officeName") if office else None,
        "resolved_benivo_office_id": office.get("officeId") if office else None,
        "job_title": candidate.get("job_title"),
        "department": candidate.get("department"),
        "location": candidate.get("location"),
        "benivo_status": candidate.get("benivo_status"),
        "is_vip": candidate.get("is_vip"),
        "policy_name": resolve_policy(candidate.get("is_vip")),
        "reason": reason,
        "selected_for_current_run": "Yes" if selected else "No",
    }


def _write_table_sheet(ws: Worksheet, rows: List[Dict[str, Any]], empty_note: Optional[str] = None) -> None:
    if not rows and empty_note:
        ws.append([empty_note])
        return

    ws.append(REQUIRED_COLUMNS)

    for row in rows:
        ws.append([row.get(column) for column in REQUIRED_COLUMNS])

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _write_summary_sheet(ws: Worksheet, summary: Dict[str, Any]) -> None:
    ws.append(["metric", "value"])

    for key, value in summary.items():
        ws.append([key, value])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30


def _fetch_refdata_if_allowed() -> Optional[Dict[str, Any]]:
    if not posting_service.allow_reference_data_calls():
        return None

    try:
        token = benivo_client.get_access_token()
        return benivo_client.get_refdata(token)
    except Exception:
        logger.exception(
            "BENIVO_ALLOW_REFERENCE_DATA_CALLS is set but fetching refdata failed; "
            "office resolution will be treated as unresolved for this report."
        )
        return None


def _posting_result_row(candidate: Dict[str, Any], result: Dict[str, Any], selected: bool) -> Dict[str, Any]:
    status = posting_service._post_log_status_from_outcome(result["outcome"])

    reasons = {
        "SUCCESS": "Candidate successfully created in Benivo.",
        "ALREADY_EXISTS": "Candidate already exists in Benivo.",
        "FAILED": result.get("error_message") or "Posting attempt failed.",
    }

    office = None
    if result.get("benivo_user_id") is not None or result.get("request_payload"):
        office = {
            "officeName": (result.get("request_payload") or {}).get("officeName"),
            "officeId": (result.get("request_payload") or {}).get("officeId"),
        }

    row = _build_row(candidate, reasons.get(status, status), office=office, selected=selected)
    row["benivo_status"] = status
    return row


def generate_reports(
    selected_candidates: List[Dict[str, Any]],
    posting_results: List[Dict[str, Any]],
    dry_run: bool,
    posting_limit: int,
) -> Path:
    """
    Build the operational Excel report from the current synchronized/classified
    data plus this run's selection and posting outcome. Must be called AFTER
    select_postable_candidates()/post_candidates() so selected_for_current_run
    and the posting-outcome counts reflect the actual current execution.
    """
    all_candidates = candidate_repository.get_all_candidates_for_report()
    terminal_eids = post_log_repository.get_terminal_post_log_application_eids()
    selected_eids = {c.get("application_eid") for c in selected_candidates}

    mobility_candidates = [c for c in all_candidates if c.get("workflow_state") == MOBILITY_WORKFLOW_STATE]

    vip_candidates = sum(1 for c in all_candidates if c.get("is_vip") is True)
    basic_candidates = len(all_candidates) - vip_candidates

    relocation_yes, relocation_no, relocation_unrecognized = [], [], []
    for candidate in mobility_candidates:
        bucket = _relocation_bucket(candidate.get("is_relocation_required"))
        {"yes": relocation_yes, "no": relocation_no, "blank_or_unrecognized": relocation_unrecognized}[bucket].append(candidate)

    relocation_yes_with_start_date = [c for c in relocation_yes if c.get("start_date")]

    # Sheet placement reads the already-persisted benivo_status directly.
    # classification_service is the single place that decides READY_TO_POST
    # vs PENDING_OFFICE_MAPPING (via the same static WORKPLACE_TO_OFFICE_NAME
    # mapping) -- this report displays that decision, it does not re-derive it.
    ready_to_post_population = [c for c in mobility_candidates if c.get("benivo_status") == "READY_TO_POST"]
    pending_office_mapping_population = [c for c in mobility_candidates if c.get("benivo_status") == "PENDING_OFFICE_MAPPING"]
    pending_missing_start_date_population = [c for c in mobility_candidates if c.get("benivo_status") == "PENDING_MISSING_START_DATE"]

    refdata = _fetch_refdata_if_allowed()

    # officeName/officeId here are display enrichment only (live refdata
    # lookup), not a readiness decision -- readiness was already decided by
    # classification_service and is reflected in ready_to_post_population.
    ready_to_post_rows: List[Dict[str, Any]] = []
    terminal_in_ready_population = 0

    for candidate in ready_to_post_population:
        application_eid = candidate.get("application_eid")

        if application_eid in terminal_eids:
            terminal_in_ready_population += 1
            continue

        selected = application_eid in selected_eids
        office = resolve_office(candidate, refdata) if refdata is not None else None
        ready_to_post_rows.append(_build_row(candidate, READY_REASON, office=office, selected=selected))

    missing_office_rows = [
        _build_row(c, MISSING_OFFICE_REASON, selected=c.get("application_eid") in selected_eids)
        for c in pending_office_mapping_population
    ]

    missing_start_date_rows = [
        _build_row(c, MISSING_START_DATE_REASON, selected=c.get("application_eid") in selected_eids)
        for c in pending_missing_start_date_population
    ]

    relocation_review_rows = [
        _build_row(c, RELOCATION_NO_REASON, selected=c.get("application_eid") in selected_eids)
        for c in relocation_no
    ] + [
        _build_row(c, RELOCATION_UNRECOGNIZED_REASON, selected=c.get("application_eid") in selected_eids)
        for c in relocation_unrecognized
    ]

    candidates_by_eid = {c.get("application_eid"): c for c in all_candidates}
    posting_result_rows = []

    for candidate, result in zip(selected_candidates, posting_results):
        if "outcome" not in result:
            continue  # dry-run preview rows have no real outcome -- excluded, per DRY_RUN_NOTE below

        full_candidate = candidates_by_eid.get(candidate.get("application_eid"), candidate)
        posting_result_rows.append(_posting_result_row(full_candidate, result, selected=True))

    posting_success = sum(1 for r in posting_result_rows if r["benivo_status"] == "SUCCESS")
    posting_already_exists = sum(1 for r in posting_result_rows if r["benivo_status"] == "ALREADY_EXISTS")
    posting_failed = sum(1 for r in posting_result_rows if r["benivo_status"] == "FAILED")

    terminal_already_processed = sum(1 for c in all_candidates if c.get("application_eid") in terminal_eids)

    summary = {
        "report_generated_at": _excel_safe(datetime.now(timezone.utc)),
        "total_current_candidates": len(all_candidates),
        "total_workflow_mobility_in_process": len(mobility_candidates),
        "vip_candidates": vip_candidates,
        "basic_candidates": basic_candidates,
        "relocation_yes": len(relocation_yes),
        "relocation_no": len(relocation_no),
        "relocation_blank_or_unrecognized": len(relocation_unrecognized),
        "relocation_yes_with_start_date": len(relocation_yes_with_start_date),
        "start_date_present": len(relocation_yes_with_start_date),
        "start_date_missing": len(relocation_yes) - len(relocation_yes_with_start_date),
        "ready_to_post": len(ready_to_post_rows),
        "pending_office_mapping": len(pending_office_mapping_population),
        "pending_missing_start_date": len(pending_missing_start_date_population),
        "relocation_requires_review": len(relocation_no) + len(relocation_unrecognized),
        "office_mapping_resolved": len(ready_to_post_rows),
        "office_mapping_unresolved": len(pending_office_mapping_population),
        "terminal_already_processed": terminal_already_processed,
        "selected_for_current_run": len(selected_eids),
        "posting_limit": posting_limit,
        "dry_run": dry_run,
        "posting_success_current_run": posting_success,
        "posting_already_exists_current_run": posting_already_exists,
        "posting_failed_current_run": posting_failed,
    }

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary_sheet(summary_ws, summary)

    _write_table_sheet(wb.create_sheet("Ready to Post"), ready_to_post_rows)
    _write_table_sheet(wb.create_sheet("Missing Start Date"), missing_start_date_rows)
    _write_table_sheet(wb.create_sheet("Relocation Field Review"), relocation_review_rows)
    _write_table_sheet(wb.create_sheet("Missing Office Mapping"), missing_office_rows)
    _write_table_sheet(
        wb.create_sheet("Posting Results"),
        posting_result_rows,
        empty_note=DRY_RUN_NOTE if dry_run else "No posting attempts were recorded in this run.",
    )

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report_path = _report_dir() / f"{REPORT_FILENAME_PREFIX}_{timestamp}.xlsx"

    wb.save(report_path)

    logger.info(
        "Report generated: %s (ready_to_post=%d, missing_start_date=%d, relocation_review=%d, missing_office_mapping=%d, posting_results=%d).",
        report_path,
        len(ready_to_post_rows),
        len(missing_start_date_rows),
        len(relocation_review_rows),
        len(missing_office_rows),
        len(posting_result_rows),
    )

    return report_path
