"""LEGACY, UNUSED by the active pipeline -- see app/services/posting_service.py.

Standalone single-candidate Benivo API connectivity test script (hardcoded
test user, not fed by real candidate data). Superseded by
posting_service.post_single_candidate(), which generalizes this same
token/refdata/create/verify flow to real candidates. Kept, not deleted, per
explicit instruction.
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook


load_dotenv()


# ============================================================
# SOLO VIENEN DEL .env ESTAS 3 VARIABLES
# ============================================================
CLIENT_ID = os.getenv("BENIVO_CLIENT_ID")
CLIENT_SECRET = os.getenv("BENIVO_CLIENT_SECRET")
GRANT_TYPE = os.getenv("BENIVO_GRANT_TYPE")
# ============================================================


# URLs FIJAS PARA UAT
TOKEN_URL = "https://externalapi.uat.benivo.com/idm/v1/Token/OAuth2"
REFDATA_URL = "https://hubapi.uat.benivo.com/v3/api/user/refdata"
CREATE_USER_URL = "https://hubapi.uat.benivo.com/v3/api/user/create"
USER_ASSIGNMENT_URL = "https://hubapi.uat.benivo.com/v3/api/user/userassignment"

# Valores de prueba
PREFERRED_POLICY = "Tier 1"
PREFERRED_OFFICE_NAME = "Colombia (Live Casino)"

# Cambia este email si quieres probar otro usuario.
TEST_EMAIL = "esteban.alvarez@arrise.com"

# Reporte en la misma carpeta del script
SCRIPT_DIR = Path(__file__).resolve().parent
REPORT_PATH = SCRIPT_DIR / "benivo_insert_report.xlsx"


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def print_json(title: str, data: Any) -> None:
    print(f"\n========== {title} ==========")
    print(json.dumps(data, indent=2, ensure_ascii=False))


def response_text(response: Optional[requests.Response]) -> str:
    if response is None:
        return ""

    try:
        return json.dumps(response.json(), ensure_ascii=False)
    except Exception:
        return response.text or ""


def validate_env() -> None:
    missing = []

    if not CLIENT_ID:
        missing.append("BENIVO_CLIENT_ID")

    if not CLIENT_SECRET:
        missing.append("BENIVO_CLIENT_SECRET")

    if not GRANT_TYPE:
        missing.append("BENIVO_GRANT_TYPE")

    if missing:
        raise RuntimeError(f"Faltan variables en .env: {', '.join(missing)}")

    if GRANT_TYPE != "client_credentials":
        raise RuntimeError("BENIVO_GRANT_TYPE debe ser exactamente: client_credentials")


def get_access_token() -> str:
    print("1) Generando token...")

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": "PostmanRuntime/7.43.0",
    }

    data = {
        "grant_type": GRANT_TYPE,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }

    response = requests.post(
        TOKEN_URL,
        headers=headers,
        data=data,
        timeout=30,
    )

    if response.status_code != 200:
        raise RuntimeError(f"No se pudo generar token: {response.status_code} - {response_text(response)}")

    payload = response.json()
    access_token = payload.get("access_token")

    if not access_token:
        raise RuntimeError(f"La respuesta del token no trajo access_token: {payload}")

    print("✅ Token generado correctamente.")
    return access_token


def get_headers(access_token: str, json_content: bool = True) -> Dict[str, str]:
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
        "User-Agent": "PostmanRuntime/7.43.0",
    }

    if json_content:
        headers["Content-Type"] = "application/json"

    return headers


def get_refdata(access_token: str) -> Dict[str, Any]:
    print("2) Consultando refdata...")

    response = requests.get(
        REFDATA_URL,
        headers=get_headers(access_token, json_content=False),
        timeout=30,
    )

    if response.status_code != 200:
        raise RuntimeError(f"No se pudo consultar refdata: {response.status_code} - {response_text(response)}")

    payload = response.json()

    if payload.get("hasError") is True:
        raise RuntimeError(f"Benivo respondió error en refdata: {payload}")

    refdata = payload.get("data")

    if not isinstance(refdata, dict):
        raise RuntimeError(f"Refdata no trae data válida: {payload}")

    print("✅ Refdata OK.")
    return refdata


def select_policy(refdata: Dict[str, Any]) -> str:
    policies = refdata.get("policies")

    if not isinstance(policies, list) or not policies:
        raise RuntimeError("No llegaron policies en refdata.")

    policy_names = [
        item.get("policy")
        for item in policies
        if isinstance(item, dict) and item.get("policy")
    ]

    if not policy_names:
        raise RuntimeError("No encontré nombres de policy en refdata.")

    if PREFERRED_POLICY in policy_names:
        return PREFERRED_POLICY

    print(f"⚠️ No encontré '{PREFERRED_POLICY}'. Usaré '{policy_names[0]}'.")
    return policy_names[0]


def select_office(refdata: Dict[str, Any]) -> Dict[str, str]:
    offices = refdata.get("offices")

    if not isinstance(offices, list) or not offices:
        raise RuntimeError("No llegaron offices en refdata.")

    for office in offices:
        if office.get("officeName") == PREFERRED_OFFICE_NAME:
            return {
                "officeId": office["id"],
                "officeName": office.get("officeName", ""),
            }

    first_office = offices[0]

    print(
        f"⚠️ No encontré '{PREFERRED_OFFICE_NAME}'. "
        f"Usaré '{first_office.get('officeName')}'."
    )

    return {
        "officeId": first_office["id"],
        "officeName": first_office.get("officeName", ""),
    }


def build_test_user(policy: str, office_id: str) -> Dict[str, Any]:
    return {
        "firstName": "Esteban",
        "lastName": "test",
        "email": TEST_EMAIL,
        "policy": policy,
        "officeId": office_id,
        "startDateOfAssignment": now_utc(),
    }


def find_user_by_email(access_token: str, email: str) -> Dict[str, Any]:
    """
    IMPORTANTE:
    Este endpoint debe recibir un OBJETO, no un array.

    Correcto:
    {"email": "test@domain.com"}

    Incorrecto:
    [{"email": "test@domain.com"}]
    """

    print(f"Buscando usuario por email: {email}")

    payload = {
        "email": email,
    }

    response = requests.post(
        USER_ASSIGNMENT_URL,
        headers=get_headers(access_token),
        json=payload,
        timeout=30,
    )

    result = {
        "found": False,
        "status_code": response.status_code,
        "email": email,
        "benivoId": None,
        "assignmentIds": [],
        "raw_response": None,
        "error": None,
    }

    try:
        body = response.json()
        result["raw_response"] = body
    except Exception:
        result["error"] = response.text
        return result

    if response.status_code != 200:
        result["error"] = response_text(response)
        return result

    if body.get("hasError") is True:
        result["error"] = response_text(response)
        return result

    data = body.get("data")

    if not isinstance(data, dict):
        return result

    user = data.get("user")

    if isinstance(user, dict):
        fetched_email = user.get("email")

        if fetched_email and fetched_email.lower() == email.lower():
            result["found"] = True
            result["benivoId"] = user.get("benivoId") or user.get("benivoID") or user.get("BenivoID")

    assignments = data.get("assignments")

    if isinstance(assignments, list):
        assignment_ids = []

        for assignment in assignments:
            if not isinstance(assignment, dict):
                continue

            assignment_id = assignment.get("assignmentId") or assignment.get("assignmentID")

            if assignment_id is not None:
                assignment_ids.append(assignment_id)

        result["assignmentIds"] = assignment_ids

    return result


def create_user(access_token: str, user_payload: Dict[str, Any]) -> Dict[str, Any]:
    print("3) Intentando crear usuario...")

    response = requests.post(
        CREATE_USER_URL,
        headers=get_headers(access_token),
        json=[user_payload],
        timeout=30,
    )

    result = {
        "success": False,
        "status_code": response.status_code,
        "created": None,
        "raw_response": None,
        "error": None,
    }

    try:
        body = response.json()
        result["raw_response"] = body
    except Exception:
        result["error"] = response.text
        return result

    if response.status_code != 200:
        result["error"] = response_text(response)
        return result

    if body.get("hasError") is True:
        result["error"] = response_text(response)
        return result

    created_rows = body.get("data")

    if not isinstance(created_rows, list) or not created_rows:
        result["error"] = "La respuesta de creación no trajo data."
        return result

    created = created_rows[0]

    missing = [
        field
        for field in ["benivoId", "assignmentId", "email"]
        if field not in created
    ]

    if missing:
        result["error"] = f"La respuesta de creación no trajo campos: {missing}"
        return result

    result["success"] = True
    result["created"] = created

    return result


def verify_created_by_ids(
    access_token: str,
    benivo_id: int,
    assignment_id: int,
    email: str,
) -> Dict[str, Any]:
    print("4) Verificando usuario por IDs...")

    payload = {
        "benivoId": benivo_id,
        "assignmentId": assignment_id,
        "email": email,
    }

    response = requests.post(
        USER_ASSIGNMENT_URL,
        headers=get_headers(access_token),
        json=payload,
        timeout=30,
    )

    result = {
        "verified": False,
        "status_code": response.status_code,
        "raw_response": None,
        "error": None,
    }

    try:
        body = response.json()
        result["raw_response"] = body
    except Exception:
        result["error"] = response.text
        return result

    if response.status_code != 200:
        result["error"] = response_text(response)
        return result

    if body.get("hasError") is True:
        result["error"] = response_text(response)
        return result

    data = body.get("data", {})
    user = data.get("user", {}) if isinstance(data, dict) else {}

    fetched_email = user.get("email") if isinstance(user, dict) else None

    if fetched_email and fetched_email.lower() == email.lower():
        result["verified"] = True

    return result


def ensure_report_file() -> None:
    if REPORT_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Benivo Insert Report"

    ws.append(
        [
            "run_at_utc",
            "status",
            "insert_attempted",
            "inserted",
            "exists_before_insert",
            "exists_after_failed_insert",
            "verified_after_insert",
            "email",
            "first_name",
            "last_name",
            "policy",
            "office_name",
            "office_id",
            "start_date_of_assignment",
            "benivo_id",
            "assignment_id",
            "error_message",
            "create_response",
            "lookup_response",
        ]
    )

    wb.save(REPORT_PATH)


def append_report(row: Dict[str, Any]) -> None:
    ensure_report_file()

    wb = load_workbook(REPORT_PATH)
    ws = wb["Benivo Insert Report"]

    ws.append(
        [
            row.get("run_at_utc"),
            row.get("status"),
            row.get("insert_attempted"),
            row.get("inserted"),
            row.get("exists_before_insert"),
            row.get("exists_after_failed_insert"),
            row.get("verified_after_insert"),
            row.get("email"),
            row.get("first_name"),
            row.get("last_name"),
            row.get("policy"),
            row.get("office_name"),
            row.get("office_id"),
            row.get("start_date_of_assignment"),
            row.get("benivo_id"),
            row.get("assignment_id"),
            row.get("error_message"),
            row.get("create_response"),
            row.get("lookup_response"),
        ]
    )

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)

    wb.save(REPORT_PATH)


def main() -> None:
    report_row = {
        "run_at_utc": now_utc(),
        "status": "FAILED",
        "insert_attempted": "NO",
        "inserted": "NO",
        "exists_before_insert": "UNKNOWN",
        "exists_after_failed_insert": "UNKNOWN",
        "verified_after_insert": "NO",
        "email": TEST_EMAIL,
        "first_name": "",
        "last_name": "",
        "policy": "",
        "office_name": "",
        "office_id": "",
        "start_date_of_assignment": "",
        "benivo_id": "",
        "assignment_id": "",
        "error_message": "",
        "create_response": "",
        "lookup_response": "",
    }

    try:
        validate_env()

        access_token = get_access_token()
        refdata = get_refdata(access_token)

        policy = select_policy(refdata)
        office = select_office(refdata)

        user_payload = build_test_user(
            policy=policy,
            office_id=office["officeId"],
        )

        report_row.update(
            {
                "email": user_payload["email"],
                "first_name": user_payload["firstName"],
                "last_name": user_payload["lastName"],
                "policy": user_payload["policy"],
                "office_name": office["officeName"],
                "office_id": user_payload["officeId"],
                "start_date_of_assignment": user_payload["startDateOfAssignment"],
            }
        )

        print_json("PAYLOAD TO INSERT", [user_payload])

        # 1. Revisar si existe antes de insertar
        before_lookup = find_user_by_email(access_token, user_payload["email"])
        report_row["lookup_response"] = json.dumps(before_lookup.get("raw_response"), ensure_ascii=False)[:5000]

        if before_lookup["found"]:
            print("\n⚠️ STATUS: ALREADY_EXISTS")
            print("El usuario ya existe. No se intenta insertar otra vez.")

            report_row.update(
                {
                    "status": "ALREADY_EXISTS",
                    "exists_before_insert": "YES",
                    "insert_attempted": "NO",
                    "inserted": "NO",
                    "exists_after_failed_insert": "NO_APPLIES",
                    "verified_after_insert": "YES",
                    "benivo_id": before_lookup.get("benivoId"),
                    "assignment_id": ",".join(map(str, before_lookup.get("assignmentIds", []))),
                }
            )

            return

        report_row["exists_before_insert"] = "NO"

        # 2. Intentar insertar
        report_row["insert_attempted"] = "YES"
        create_result = create_user(access_token, user_payload)
        report_row["create_response"] = json.dumps(create_result.get("raw_response"), ensure_ascii=False)[:5000]

        # 3. Si NO pudo insertar, igual buscar si existe
        if not create_result["success"]:
            print("\n⚠️ No se pudo insertar. Voy a revisar si existe por email...")

            after_failed_lookup = find_user_by_email(access_token, user_payload["email"])
            report_row["lookup_response"] = json.dumps(after_failed_lookup.get("raw_response"), ensure_ascii=False)[:5000]

            if after_failed_lookup["found"]:
                print("\n⚠️ STATUS: INSERT_FAILED_BUT_EXISTS")
                print("Falló el insert, pero el usuario sí existe en Benivo.")

                report_row.update(
                    {
                        "status": "INSERT_FAILED_BUT_EXISTS",
                        "inserted": "NO",
                        "exists_after_failed_insert": "YES",
                        "verified_after_insert": "YES",
                        "benivo_id": after_failed_lookup.get("benivoId"),
                        "assignment_id": ",".join(map(str, after_failed_lookup.get("assignmentIds", []))),
                        "error_message": create_result.get("error"),
                    }
                )

                return

            print("\n❌ STATUS: INSERT_FAILED_NOT_FOUND")
            print("Falló el insert y el usuario no aparece por email.")

            report_row.update(
                {
                    "status": "INSERT_FAILED_NOT_FOUND",
                    "inserted": "NO",
                    "exists_after_failed_insert": "NO",
                    "verified_after_insert": "NO",
                    "error_message": create_result.get("error"),
                }
            )

            return

        # 4. Si insertó, verificar por IDs y por email
        created = create_result["created"]

        print_json("CREATE RESPONSE", created)

        benivo_id = int(created["benivoId"])
        assignment_id = int(created["assignmentId"])
        email = created["email"]

        report_row.update(
            {
                "inserted": "YES",
                "benivo_id": benivo_id,
                "assignment_id": assignment_id,
            }
        )

        verify_by_ids = verify_created_by_ids(
            access_token=access_token,
            benivo_id=benivo_id,
            assignment_id=assignment_id,
            email=email,
        )

        verify_by_email = find_user_by_email(access_token, email)

        if verify_by_ids["verified"] or verify_by_email["found"]:
            print("\n✅ STATUS: INSERTED_AND_VERIFIED")
            print("Usuario insertado y verificado correctamente.")

            report_row.update(
                {
                    "status": "INSERTED_AND_VERIFIED",
                    "verified_after_insert": "YES",
                    "lookup_response": json.dumps(verify_by_email.get("raw_response"), ensure_ascii=False)[:5000],
                }
            )

            return

        print("\n⚠️ STATUS: INSERTED_BUT_NOT_VERIFIED")
        print("Benivo respondió creación OK, pero la búsqueda posterior no confirmó el usuario.")

        report_row.update(
            {
                "status": "INSERTED_BUT_NOT_VERIFIED",
                "verified_after_insert": "NO",
                "lookup_response": json.dumps(verify_by_email.get("raw_response"), ensure_ascii=False)[:5000],
            }
        )

    except Exception as exc:
        print(f"\n❌ ERROR GENERAL: {str(exc)}")

        report_row.update(
            {
                "status": "FAILED",
                "error_message": str(exc),
            }
        )

    finally:
        append_report(report_row)

        print("\n📄 Reporte actualizado en:")
        print(REPORT_PATH)


if __name__ == "__main__":
    main()